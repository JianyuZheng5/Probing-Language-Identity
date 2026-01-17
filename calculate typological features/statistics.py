# coding=gbk#coding:utf-8
#导库
import openpyxl
from openpyxl import Workbook
import argparse
import os

#write函数:把各语言类型学特征的结果写入excel表格中
def write(args):
    result = dict()
    with open(args.filename, 'r', encoding='utf-8') as fp:  #args.filename为存储数据结果的文件
        lines = fp.readlines()

    #统计结果
    for line in lines:
        tokens = line.strip().split()
        language = tokens[0].split(':')[-1]

        result[language] = result.get(language, dict())

        feature_index = int(tokens[1].split(':')[-1])
        value = float(tokens[3].split(':')[-1])
        maximum = result[language].get(feature_index, -0.01)
        if value >maximum:
            result[language][feature_index] =value

    #创建excel文件
    path = './layer '+str(args.layer)
    if not os.path.exists(path):
        os.makedirs(path)
    fn = path +'/layer'+str(args.layer)+'_result_'+args.model+'.xlsx'
    wb = Workbook()
    wb.save(fn)
    wb = openpyxl.load_workbook(fn)
    ws = wb.worksheets[0]
    #写入特征所在的列
    for i in range(len(args.features)):
        _=ws.cell(row=i+2, column=1, value= args.features[i])
    #写入语种所在的行
    for i in range(len(args.languages)):
        _=ws.cell(row=1, column=2+i, value= args.languages[i])
    #写入统计结果
    for j in range(len(args.languages)):
        language = args.languages[j]
        if language == "Greek (Modern)":
            language = 'Greek(Modern)'
        indexs = list(result[language].keys())
        for i in range(len(args.features)):
            if i in indexs:
                _ = ws.cell(row=i+2, column=j+2, value=result[language][i])
            else:
                _ = ws.cell(row=i+2, column=j+2, value='/')
    wb.save(fn)  #保存
